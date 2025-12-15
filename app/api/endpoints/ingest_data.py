from fastapi import APIRouter, UploadFile, Depends, HTTPException
from sqlalchemy.orm import Session
from app.db import SessionLocal
from app.db.crud import bulk_insert_lighting_posts, get_existing_lighting_post_ids
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from datetime import datetime
from openpyxl.utils.datetime import from_excel

ingest_router = APIRouter(prefix="/lighting-data", tags=["lighting-posts"])
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def process_date(value):
    if isinstance(value, int):  # Data no formato numérico do Excel
        return from_excel(value).date()
    elif isinstance(value, str):  # Data no formato string (ex: "12/05/2023")
        try:
            return datetime.strptime(value, "%d/%m/%Y").date()
        except ValueError:
            return None
    return None

@ingest_router.post("/upload")
async def upload_lighting_data(file: UploadFile, db: Session = Depends(get_db)):
    duplicated_posts = []
    already_in_db = []
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Arquivo inválido. Apenas arquivos .xlsx são aceitos.")

    try:
        workbook = load_workbook(file.file)
        sheet = workbook.active
        posts_data = []
        ids_in_file = set()
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            try:
                row_id = row[5]

                # detectar duplicado dentro do arquivo
                if row_id in ids_in_file:
                    print(f"Registro duplicado no arquivo: ID {row_id}.")
                    duplicated_posts.append({"id": str(row_id), "address": row[0]})
                    continue
                ids_in_file.add(row_id)

                installation_date = process_date(row[7])
                last_maintenance = process_date(row[8])

                post_data = {
                    "id": row_id,
                    "address": row[0],
                    "latitude": row[2],
                    "longitude": row[1],
                    "lamp_type": row[3],
                    "wattage": row[4],
                    "voltage": int(str(row[10]).replace("V", "")),
                    "needs_repair": row[9].lower() == "sim",
                    "last_maintenance": last_maintenance,
                    "installation_date": installation_date,
                    "hours_in_operation": row[9]
                }

                posts_data.append(post_data)
            except Exception as e:
                print(f"Erro ao processar linha {row_index}: {e}")
                continue

        ids_to_check = [p["id"] for p in posts_data if p["id"] is not None]
        existing_ids_set = set(get_existing_lighting_post_ids(db, ids_to_check))
        if existing_ids_set:
            for ex in existing_ids_set:
                already_in_db.append({"id": str(ex)})
            # filtra posts_data para inserir apenas os que não existem
            posts_to_insert = [p for p in posts_data if p["id"] not in existing_ids_set]
        else:
            posts_to_insert = posts_data

        try:
            total_inserted = bulk_insert_lighting_posts(db, posts_to_insert)
        except IntegrityError as e:
            print("IntegrityError no bulk insert, tentando inserts individuais:", e)
            total_inserted = 0
        except Exception as e:
            print("Erro durante o Bulk Insert:", e)
            raise HTTPException(status_code=500, detail="Erro durante o Bulk Insert")

        return {
            "message": f"Inseridos {total_inserted} registros com sucesso.",
            "duplicate_values_in_file": duplicated_posts,
            "duplicate_values_in_db": already_in_db
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))